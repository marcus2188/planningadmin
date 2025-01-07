import json
import re, subprocess, time
from PyQt6.QtWidgets import QFileDialog, QWidget, QMessageBox
from PyQt6.QtCore import QSettings, Qt
from data import single_source_of_truth
from datetime import datetime, timedelta
import openpyxl as op
from openpyxl.utils.dataframe import dataframe_to_rows
import pandas as pd
import numpy as np
import traceback
import os


def launch_message_box(mw, titlemsg, content):
    """Summons the useful error message box from anywhere you want"""

    # Prompt message box
    msgbox = QMessageBox()
    msgbox.setWindowTitle(titlemsg)
    msgbox.setText(content)
    msgbox.exec()

    # Reenable main window and hide progressbar
    mw.central_widget.setDisabled(False)
    mw.progress_window.hide()


def gr_add_browsed_file(mw, filetype: str):
    """In generate ecd page, launches finder window and add browsed file into display
    Follows list_of_files_to_browse index order (ind_of_file)
    """
    filepath, _ = QFileDialog.getOpenFileName(
        QWidget(), f"Open {filetype} File", "", "Excel files (*.xlsm *.xls *.xlsx)"
    )

    # Set the correct textfield
    mw.gr_browse_text_fields[mw.list_of_files_to_browse.index(filetype)].setPlainText(
        filepath
    )


def gr_clear_all_added_files(mw):
    """In generate ecd page, clears all text fields"""
    for tf in mw.gr_browse_text_fields:
        tf.setPlainText("")


def gr_auto_find_all_files(mw):
    """In generate ecd page, dig up these files using name regex pattern matching"""
    perma_wk_dir = mw.ss_perma_working_dir_text_field.toPlainText()

    # Check if perma working dir is available
    if perma_wk_dir == "":
        launch_message_box(
            mw,
            "No permanent directory (PWD)",
            "You have not set a permanent working directory to auto find files from, inside the settings page.",
        )
        return

    # Walk through perma working dir and do string matching
    for found_file in sorted(
        os.scandir(perma_wk_dir), key=lambda f: os.path.getctime(f)
    ):
        if found_file.is_file() and found_file.name.upper().endswith(".XLSX"):
            for x, v in single_source_of_truth.mapped_sap_file_to_filetype.items():
                if x in found_file.name.upper():
                    index = mw.list_of_files_to_browse.index(v)
                    mw.gr_browse_text_fields[index].setPlainText(
                        os.path.join(perma_wk_dir, found_file.name)
                    )


def ss_change_perma_working_dir(mw):
    """In settings page, change perma wkdir textfield"""
    # Launch finder window
    folderpath = QFileDialog.getExistingDirectory(
        QWidget(), "Choose your permanent working directory to find and save to"
    )

    # Change textfield and .ini settings file
    settings = QSettings("greatsettings.ini", QSettings.Format.IniFormat)
    mw.ss_perma_working_dir_text_field.setPlainText(folderpath)
    settings.setValue(
        "perma_working_dir", mw.ss_perma_working_dir_text_field.toPlainText()
    )


def ss_clear_perma_working_dir(mw):
    """In settings page, delete perma wkdir textfield"""
    settings = QSettings("greatsettings.ini", QSettings.Format.IniFormat)
    mw.ss_perma_working_dir_text_field.setPlainText("")
    settings.setValue(
        "perma_working_dir", mw.ss_perma_working_dir_text_field.toPlainText()
    )


def ss_change_default_ecdgen_name(mw):
    """In settings page, change ecdgen name and in settings"""
    settings = QSettings("greatsettings.ini", QSettings.Format.IniFormat)
    settings.setValue("default_ecdgen_name", mw.ss_ecdgen_filename_text_field.text())


def ss_change_ascent_color(mw):
    """In settings page, change ascent color and in settings"""
    settings = QSettings("greatsettings.ini", QSettings.Format.IniFormat)

    # Amongst the ascent color radio buttons, check which is currently set
    for rb in mw.ss_accent_radio_buttons:
        if rb.isChecked():
            settings.setValue("chosen_ascent_color", rb.radio_text)

            # Change bg colour of sidebar and each pageframe
            mw.central_side_bar.set_color(
                rbg_tuple=mw.mapped_rgb_tuples_to_colors[rb.radio_text]["sidebar"]
            )
            mw.generate_rupture_frame.set_color(
                rbg_tuple=mw.mapped_rgb_tuples_to_colors[rb.radio_text]["pageframe"]
            )
            mw.settings_frame.set_color(
                rbg_tuple=mw.mapped_rgb_tuples_to_colors[rb.radio_text]["pageframe"]
            )


def vb_add_browsed_vbs_file(mw, filetype: str):
    """In generate ecd page, launches finder window and add browsed file into display
    Follows list_of_files_to_browse index order (ind_of_file)
    """
    filepath, _ = QFileDialog.getOpenFileName(
        QWidget(),
        f"Please select a {filetype} file from your system",
        "",
        "VBScript files (*.vbs)",
    )

    # Add this path into vbs_path_history list, and deduplicate list while preserving order
    mw.vbs_path_history.append(filepath)
    mw.vbs_path_history = list(dict.fromkeys(mw.vbs_path_history))

    # Refresh vbs list widget history
    mw.vbs_list_widget.clear()
    mw.vbs_list_widget.addItems(mw.vbs_path_history)

    # select the current browsed file/same file
    mw.vbs_list_widget.setCurrentRow(mw.vbs_path_history.index(filepath))

    # Save history list into settings
    settings = QSettings("greatsettings.ini", QSettings.Format.IniFormat)
    settings.setValue("vbs_path_history", json.dumps(mw.vbs_path_history))


def vb_clear_all_vbs_history_list(mw):
    """In vbsloader page, clear the vbs history list of paths"""
    mw.vbs_list_widget.clear()
    mw.vbs_path_history = []
    
    # Save history list into settings
    settings = QSettings("greatsettings.ini", QSettings.Format.IniFormat)
    settings.setValue("vbs_path_history", json.dumps(mw.vbs_path_history))


def vb_execute_vbs_script(mw):
    """Actually runs the vbs script on run vbs button press"""

    # First get current qlistwidget selection and check if there even is one selected
    current_item = mw.vbs_list_widget.currentItem()
    if not current_item or current_item.text().replace(" ", "") == "":
        launch_message_box(
            mw,
            "No vbs / blank selection detected",
            "You have not selected any vbs to run / blank selection. Please click on one of the vbs in the list, or browse for a vbs script",
        )
        return

    # Get the current selection vbs path and execute in bash shell
    literal_path_str = os.path.normpath(current_item.text())

    # Verify if vbs path is valid
    if not os.path.exists(literal_path_str):
        launch_message_box(
            mw,
            "Invalid vbs location",
            "The selected vbs file cannot be found. Please browse for a new one or select another",
        )
        return

    # Start static progress bar and the script run
    mw.progress_window.displayPlease()
    mw.central_widget.setDisabled(True)
    mw.progress_window.updateProgressBar(0)
    try:
        startup_info = subprocess.STARTUPINFO()
        startup_info.dwFlags |= subprocess.STARTF_USESHOWWINDOW
        startup_info.wShowWindow = subprocess.SW_HIDE

        subprocess.run(["cscript", "//nologo", literal_path_str], check=True, startupinfo=startup_info)

    except Exception as eee:
        exception_name = type(eee).__name__
        exception_description = str(eee)
        launch_message_box(
            mw,
            "An Error has occurred, please report incident to Marcus thanks",
            f"{exception_name},{exception_description}",
        )
        traceback.print_exc()
    mw.central_widget.setDisabled(False)
    mw.progress_window.updateInternalText("VBS script executed.")
    mw.progress_window.updateProgressBar(100)
    mw.progress_window.unhideOkButton()

def run_rupture_algorithm(mw, myws, dict_of_pd_excel_dfs):
    """
    Main rupture algorithm happens here.
    """
    
    # Build master dataframe template through vertical concat, streamline final formatting first
    master_df_columns_template = [
        "Customer",
        "Order Type",
        "Order No",
        "Material Number",
        "Material Description",
        "Plant",
        "Qty",
        "Acc Qty",
        "Status",
        "First Date",
        "ECD Date",
        "Filling Date",
        "Delay",
        "Cat",
        "Remarks"
    ]
    
    ## Pad raw cols into master cols template strictly as above
    master_list = []
    
    # SOH - Add in rows from Opening SOH doc
    prev_month_last_date = (datetime.now().replace(day=1) - timedelta(days=1)).strftime("%d/%m/%Y")
    for r in dict_of_pd_excel_dfs["Opening SOH"].values:
        master_list.append(
            [
                None,               ## Customer keep blank for now
                "SOH",              ## Order Type is always "SOH"
                None,               ## Order No keep blank for now
                r[1],               ## Material Number
                r[2],               ## Material Desc
                r[0],               ## Plant
                r[3],               ## Qty
                None,               ## Acc Qty keep blank for now, need to post-calculate
                None,               ## status, keep blank for now
                prev_month_last_date,       ## First Date, use prev_month_last_date for actual. hardcoding for now
                None,              ## ECD Date leave blank for now
                None,              ## Filling Date leave blank for now
                None,              ## Delay leave blank for now
                None,              ## Cat leave blank for now
                None               ## Remarks leave blank for now
            ]
        )
    
    # ZSO - Add in rows from ZC32 doc after filtered applied
    zc32_df = dict_of_pd_excel_dfs["ZC32"]
    zc32_df = zc32_df[
        # (zc32_df['Source'] == 'SG') &
        (zc32_df['Delivery Date'].dt.year == datetime.now().year) &
        (zc32_df['Reason For Rejection'].isnull())
    ]
    for r in zc32_df.values:
        master_list.append(
            [
                r[5],               ## Customer keep blank for now
                "ZSO",              ## Order Type is always "ZSO"
                r[2],               ## Order No keep blank for now
                r[7],               ## Material Number
                r[8],               ## Material Desc
                r[13],               ## Plant
                r[15] * -1,          ## Qty
                None,               ## Acc Qty keep blank for now, need to post-calculate
                None,               ## status, keep blank for now
                r[11].strftime('%d/%m/%Y'),          ## First Date, use prev_month_last_date for actual. hardcoding for now
                None,              ## ECD Date leave blank for now
                None,              ## Filling Date leave blank for now
                None,              ## Delay leave blank for now
                None,              ## Cat leave blank for now
                None               ## Remarks leave blank for now
            ]
        )
    
    # ZPO3 - Add in rows from Prod Vol doc after filtered applied
    prod_vol_df = dict_of_pd_excel_dfs["Prod Vol"]
    prod_vol_df = prod_vol_df[prod_vol_df['Unit of measure (=GMEIN)'] == 'ST']
    for r in prod_vol_df.values:
        master_list.append(
            [
                None,               ## Customer keep blank for now
                "ZPO3",              ## Order Type is always "ZSO"
                r[1],               ## Order No keep blank for now
                r[3],               ## Material Number
                r[4],               ## Material Desc
                r[0],               ## Plant
                r[7],              ## Qty
                None,               ## Acc Qty keep blank for now, need to post-calculate
                None,               ## status, keep blank for now
                (r[13] + timedelta(days=3)).strftime('%d/%m/%Y'),          ## First Date, +3 days for ZPO3
                None,              ## ECD Date leave blank for now
                None,              ## Filling Date leave blank for now
                None,              ## Delay leave blank for now
                None,              ## Cat leave blank for now
                None               ## Remarks leave blank for now
            ]
        )

    # DN - Add in rows from DN ZC48 doc after filtered applied
    dn_df = dict_of_pd_excel_dfs["DN"]
    dn_df = dn_df[
        (dn_df['Unit'] == 'ST') &
        (dn_df['Actual GI date'].dt.year == datetime.now().year)
        # (dn_df['Unnamed: 14'] == 'SG')
    ]
    for r in dn_df.values:
        master_list.append(
            [
                r[1],               ## Customer keep blank for now
                "DN",              ## Order Type is always "DN"
                r[9],               ## Order No keep blank for now
                r[12],               ## Material Number
                r[13],               ## Material Desc
                r[3],               ## Plant
                r[17] * -1,          ## Qty
                None,               ## Acc Qty keep blank for now, need to post-calculate
                None,               ## status, keep blank for now
                r[10].strftime('%d/%m/%Y'),          ## First Date, use prev_month_last_date for actual. hardcoding for now
                None,              ## ECD Date leave blank for now
                None,              ## Filling Date leave blank for now
                None,              ## Delay leave blank for now
                None,              ## Cat leave blank for now
                None               ## Remarks leave blank for now
            ]
        )
    
    # STO - From the H978 doc
    sto_df = dict_of_pd_excel_dfs["Stock Transfer"]
    # dn_df = dn_df[
    #     (dn_df['Unit'] == 'ST') &
    #     (dn_df['Actual GI date'].dt.year == datetime.now().year) &
    #     (dn_df['Unnamed: 14'] == 'SG')
    # ]
    for r in sto_df.values:
        master_list.append(
            [
                "",               ## Customer keep blank for now
                "STO",              ## Order Type is always "DN"
                "",               ## Order No keep blank for now
                r[1],               ## Material Number
                r[2],               ## Material Desc
                r[0],               ## Plant
                r[5],          ## Qty
                None,               ## Acc Qty keep blank for now, need to post-calculate
                None,               ## status, keep blank for now
                r[3].strftime('%d/%m/%Y'),          ## First Date, use prev_month_last_date for actual. hardcoding for now
                None,              ## ECD Date leave blank for now
                None,              ## Filling Date leave blank for now
                None,              ## Delay leave blank for now
                None,              ## Cat leave blank for now
                None               ## Remarks leave blank for now
            ]
        )
        
    # Put data into final df & sort by matcode, then firstdate then order type
    master_df = pd.DataFrame(master_list, columns=master_df_columns_template)
    master_df['tempdate'] = pd.to_datetime(master_df['First Date'], format='%d/%m/%Y')
    master_df = master_df.sort_values(
        by=['Material Number', 'tempdate', 'Order Type'],
        ascending=[True, True, True],
        key=lambda x: x.map({'SOH': 1, 'ZPO3': 2, 'DN': 3, 'STO': 4, 'ZSO': 5}) if x.name == 'Order Type' else x
    )
    master_df = master_df.drop(columns=['tempdate'])
    
    # Calculate Accum Qty
    master_df['Acc Qty'] = master_df.groupby('Material Number')['Qty'].cumsum()

    # Apply rupture status checks
    master_df['Status'] = master_df['Acc Qty'].apply(lambda x: 'Ok' if x > 0 else 'Nok')
    
    ## Directly write master df into excel workbook sheet
    for r in dataframe_to_rows(master_df, index=False):
        myws.append(r)
        
    return myws, None, master_df

def run_ecd_algorithm(mw, myws, dict_of_pd_excel_dfs, mode):
    """
    Where the main ecd allocation algorithm happens.
    Mode --> Either "main" or "freestock"
    """

    # Data preprocessing: Preprocess and raw data first before building master df
    ## mb52 SOH df: Convert qty column in rows with 'packed' in mat desc col into drum/ST
    dict_of_pd_excel_dfs["SOH"].loc[
        dict_of_pd_excel_dfs["SOH"]["Material Description"].str.contains("PACKED"),
        "Unrestricted",
    ] = (
        (dict_of_pd_excel_dfs["SOH"]["Unrestricted"] / 208).round().astype(int)
    )

    # Build master dataframe template through vertical concat, streamline final formatting first
    master_df_columns_template = [
        "Material Number",
        "Material Description",
        "Primary Key",
        "Sales/Delivery Number",
        "Name",
        "Type",
        "Item",
        "First Delivery Date",
        "Open Qty",
        # Next 6 cols added in only after forming the above 9 columns
        "Counter",
        "Reallocation",
        "Pending",
        "ECD",
        "Change",
        "Comments",
    ]

    ## Pad raw cols into master cols template strictly as above
    master_list = []

    ### SO (sales order): Sales taken out of inventory to sell to customers, only take unrejected ones
    so_df = dict_of_pd_excel_dfs["SO"]
    filtered_so_df = so_df[so_df["Reason For Rejection"].isnull()]
    for r in filtered_so_df.values:
        master_list.append(
            [
                r[0],  ### mat and matcode keep as standard
                r[1],
                str(r[6])
                + str(r[8]),  ### primary key formed concating sales doc + item as strs
                r[6],
                r[4],
                "SO",
                r[8],
                r[3].strftime(
                    "%d/%m/%Y"
                ),  ### Reformat timestamp into sg date format dd/mm/yyyy
                r[2] * -1,
            ]
        )

    ### SOH (stock on hand): Amt currently in the inventory
    for r in dict_of_pd_excel_dfs["SOH"].values:
        master_list.append(
            [
                r[0],  ### mat and matcode keep as standard
                r[1],
                "",  ### No sales doc nor item so no primary key
                np.nan,  ### No sales doc
                np.nan,  ### No name
                "SOH",
                np.nan,  ### No item no
                "01/12/2019",  ### Possible hardcoded date str
                int(r[2]),  ### Convert float to int via casting cos no decimal parts
            ]
        )

    ### Conrel (Special orders for special clients): Sales orders that go to special clients?
    if not dict_of_pd_excel_dfs["Conrel"].empty:
        for r in (
            dict_of_pd_excel_dfs["Conrel"]
            .drop(dict_of_pd_excel_dfs["Conrel"].tail(1).index)
            .values
        ):
            master_list.append(
                [
                    r[3],  ### mat and matcode keep as standard
                    r[4],
                    "",  ### No sales doc nor item so no primary key
                    np.nan,  ### No sales doc
                    np.nan,  ### No name
                    "Conrel",
                    np.nan,  ### No item no
                    r[8].strftime("%d/%m/%Y"),  ### Reformat dev creation date
                    r[5]
                    * -1,  ### negative open quantity cos its taking away from inventory
                ]
            )

    ### DN (Delivery Notice sales orders): Sales orders that are preplanned ahead of delivery
    for r in dict_of_pd_excel_dfs["DN"].values:
        master_list.append(
            [
                r[2],  ### mat and matcode keep as standard
                r[3],
                "",  ### No sales doc nor item so no primary key
                r[0],
                r[5],
                "DN",
                np.nan,  ### No item no
                "02/12/2019",  ### Possible hardcoded date str
                r[4]
                * -1,  ### negative open quantity cos its taking away from inventory
            ]
        )

    ### OISL (Replenish Inventory)
    if mode == "main":
        if not dict_of_pd_excel_dfs["OISL"].empty:
            for r in dict_of_pd_excel_dfs["OISL"].values:
                master_list.append(
                    [
                        r[0],  ### mat and matcode keep as standard
                        r[1],
                        "",  ### No sales doc nor item so no primary key
                        np.nan,
                        np.nan,
                        "OISL",
                        np.nan,  ### No item no
                        r[3].strftime("%d/%m/%Y"),
                        r[2],
                    ]
                )

    # Create master dataframe, sort by increasing mat + earlier to later date after temp wrap casting datestr
    master_df = pd.DataFrame(master_list, columns=master_df_columns_template[:9])
    master_df["First Delivery Date"] = pd.to_datetime(
        master_df["First Delivery Date"], format="%d/%m/%Y"
    )
    master_df.sort_values(
        by=["Material Number", "First Delivery Date"], ascending=True, inplace=True
    )
    master_df["First Delivery Date"] = master_df["First Delivery Date"].dt.strftime(
        "%d/%m/%Y"
    )

    # Add in the 6 additional columns into master df
    (
        master_df["Counter"],
        master_df["Reallocation"],
        master_df["Pending"],
        master_df["ECD"],
        master_df["Change"],
        master_df["Comments"],
    ) = (0, 0, 0, "", "-", "")

    ## Directly put master df into excel workbook
    for r in dataframe_to_rows(master_df, index=False):
        myws.append(r)

    ## Init the intermediate data structures to store computation products
    interim_per_mat_dict = {
        mat: {
            "balance": 0,  # Stores overall pending amt (all outflows from same mat that are not allocated) due to lack of intakes
            "intakes": [],  # Stores available or incoming stock (SOH, OISL etc): [[openqty, firstdevdate, type], [..], ...]
            "overshot": 0,  # Tracks
        }
        for mat in np.unique(master_df[["Material Number"]].values.tolist())
    }

    ## Add into interim ds, all positives intakes (open qty > 0) in master df
    for sapcode in interim_per_mat_dict.keys():
        for r in range(
            2, master_df.shape[0] + 2
        ):  ### difference in starting indices between excel and pandas df is +2, excel start from row 1 and row 1 is header
            if sapcode == myws.cell(row=r, column=1).value:
                if myws.cell(row=r, column=9).value > 0:
                    interim_per_mat_dict[sapcode]["intakes"].append(
                        [
                            myws.cell(row=r, column=9).value,  ### open qty
                            myws.cell(row=r, column=8).value,  ### first dev date
                            myws.cell(row=r, column=6).value,  ### type
                        ]
                    )

    ## Fill in all 6 addition column values
    outflow, open_qty_cuml_sum = 0, 0
    for sapcode in interim_per_mat_dict.keys():
        open_qty_cuml_sum = 0  ### Reset every sapcode cycle of cos
        for r in range(
            2, master_df.shape[0] + 2
        ):  #### Find all outflows and allocate to all of them
            if sapcode == myws.cell(row=r, column=1).value:
                ### Calculate open_qty_sum by cumulatively summing all open_qty in same mat/sapcode
                open_qty_cuml_sum += myws.cell(row=r, column=9).value
                myws.cell(
                    row=r, column=10
                ).value = open_qty_cuml_sum  ### Put open_qty_cuml_sum into 'counter' col in excel

                ### Allocate the intakes to all negative open qtys (outflows) cumulatively top down order, for each mat/sapcode
                if (
                    myws.cell(row=r, column=9).value < 0
                ):  ### if openqty negative, means there is outflows to be allocated to
                    outflow = abs(
                        myws.cell(row=r, column=9).value
                    )  ### get its abs as the outflow
                    strtoadd = ""
                    for intake_ind in range(
                        len(interim_per_mat_dict[sapcode]["intakes"])
                    ):  ### Iterate thru each intake
                        intake_amount = interim_per_mat_dict[sapcode]["intakes"][
                            intake_ind
                        ][
                            0
                        ]  ### indice 0 cos rmb list in list
                        intake_date = interim_per_mat_dict[sapcode]["intakes"][
                            intake_ind
                        ][1]
                        intake_type = interim_per_mat_dict[sapcode]["intakes"][
                            intake_ind
                        ][2]

                        #### If intake amount is not zero and outflow is greater than intake amt
                        if intake_amount != 0 and outflow > intake_amount:
                            outflow -= intake_amount  #### Allocate entire intake for this outflow, subtract from outflow
                            if (
                                intake_type == "OISL"
                            ):  #### If OISL intake, then add in like normal
                                strtoadd = (
                                    strtoadd
                                    + str(intake_date)
                                    + ": "
                                    + str(int(intake_amount))
                                    + ", "
                                )  #### Add delivery date and intake amt to ecd
                            else:
                                if (
                                    "SOH" in strtoadd
                                ):  #### If SOH exists in ecd, update its value by adding in the current intake amt
                                    strtoadd = re.sub(
                                        r"SOH\: \d+, ",
                                        "SOH: "
                                        + str(
                                            int(
                                                (strtoadd.split(", ")[0]).split(": ")[1]
                                            )
                                            + intake_amount
                                        )
                                        + ", ",
                                        strtoadd,
                                    )
                                else:  #### If SOH not in, create it in ecd
                                    strtoadd = (
                                        strtoadd
                                        + "SOH"
                                        + ": "
                                        + str(int(intake_amount))
                                        + ", "
                                    )

                            ##### Zero the intake amount and continue iterating thru other intakes to finish allocating the outflow
                            interim_per_mat_dict[sapcode]["intakes"][intake_ind][0] = 0

                        #### If that outflow less than that single intake
                        elif outflow <= intake_amount:
                            interim_per_mat_dict[sapcode]["intakes"][intake_ind][
                                0
                            ] -= outflow  ##### Allocate some of the intake to outflow, must use direct reference cos of python iterative inplace problems
                            if intake_type == "OISL":
                                strtoadd = (
                                    strtoadd
                                    + str(intake_date)
                                    + ": "
                                    + str(int(outflow))
                                    + ", "
                                )
                            else:
                                if "SOH" in strtoadd:
                                    strtoadd = re.sub(
                                        r"SOH\: \d+, ",
                                        "SOH: "
                                        + str(
                                            int(
                                                (strtoadd.split(", ")[0]).split(": ")[1]
                                            )
                                            + outflow
                                        )
                                        + ", ",
                                        strtoadd,
                                    )
                                else:
                                    strtoadd = (
                                        strtoadd + "SOH" + ": " + str(int(outflow)) + ", "
                                    )

                            ##### Stop allocating intakes to that single outflow since outflow is all allocated finish
                            outflow = 0
                            break

                    ### Means overall for that mat/sapcode, not enough intake to allocate to all outflows
                    if outflow != 0:
                        myws.cell(
                            row=r, column=12
                        ).value = outflow  ### Write remaining outstanding outflow to 'pending' col
                    interim_per_mat_dict[sapcode][
                        "balance"
                    ] -= outflow  ### Store balance values

                    ### Add the ecd string in finally, together with a pending entry if the cumulative excel pending value is +ve
                    if (
                        myws.cell(row=r, column=12).value
                        and myws.cell(row=r, column=12).value > 0
                    ):
                        myws.cell(row=r, column=13).value = (
                            strtoadd
                            + "Pending: "
                            + str(int(myws.cell(row=r, column=12).value))
                            + ", "
                        )
                    else:
                        myws.cell(row=r, column=13).value = strtoadd

    ## Next we recompute balance for each sapcode by adding up leftover intake amounts
    for sapcode in interim_per_mat_dict.keys():
        interim_per_mat_dict[sapcode]["balance"] += sum(
            [px[0] for px in interim_per_mat_dict[sapcode]["intakes"]]
        )

    ## Highlight outstanding ecd numbers red
    for r in range(2, master_df.shape[0] + 2):
        if myws.cell(row=r, column=12).value > 0:
            myws.cell(row=r, column=12).font = op.styles.Font(color="FF0000")

    ## Remove all the SOH = 0s and date = 0s from ecd columns
    for r in range(2, master_df.shape[0] + 2):
        if myws.cell(row=r, column=12).value == 0:
            myws.cell(row=r, column=12).value = ""

    ## Check for not end with SG, if so add any ECD dates by 14 days
    if mw.gr_14days_add_checkbox.isChecked():  ### Now this is default checked
        for r in range(2, master_df.shape[0] + 2):
            ### If mat desc ends with "SG" and mat/sapcode between 100000 to 299999
            if (
                (myws.cell(row=r, column=2).value.endswith("SG") == False)
                and (int(myws.cell(row=r, column=1).value) <= 299999)
                and (int(myws.cell(row=r, column=1).value) >= 100000)
            ):
                ### Add 14 days via converting to datetime and adding timedelta and back into regex matched date in excel ecd col
                def add14days(match):
                    return datetime.strftime(
                        (
                            datetime.strptime(match.group(0), "%d/%m/%Y")
                            + timedelta(days=14)
                        ),
                        "%d/%m/%Y",
                    )

                myws.cell(row=r, column=13).value = re.sub(
                    r"\d+/\d+/\d+", add14days, myws.cell(row=r, column=13).value
                )
    return myws, interim_per_mat_dict, master_df


def read_in_backend_excel(mw, excelfilepath, sheetname, usecolrange):
    """Reads in and validates a backend excel file db, returns ready pandas df"""

    # Check if excelfilepath file exist in backend
    if not os.path.exists(excelfilepath):
        launch_message_box(
            mw,
            "Fatal Internal Error",
            f"An internal backend file {excelfilepath} dependency is not found, please rectify and ensure it exists",
        )
        return

    # Read into pandas df
    backend_df = pd.read_excel(
        excelfilepath,
        sheet_name=sheetname,
        usecols=range(usecolrange),
    )

    # Verify integrity (Exact column names and order + count must > 0)
    correct_columns = single_source_of_truth.ecd_file_correct_columns[excelfilepath]
    if not list(backend_df.columns) == correct_columns:
        launch_message_box(
            mw,
            "Fatal Internal Error",
            f"An internal backend file {excelfilepath} dependency is in an invalid format, please make sure the format follows the original one",
        )
        return
    if backend_df.empty:
        launch_message_box(
            mw,
            "Fatal Internal Error",
            f"An internal backend file {excelfilepath} dependency has zero data, unable to continue. Please rectify",
        )
        return
    return backend_df


def fill_sheet_with_balance_df(mw, myws, interim_per_mat_dict, master_df, mode):
    """Fills up an excel worksheet with a balance df"""

    # Create a balance_df containing sapcode and balance retrieved from interim_per_mat_dict
    balance_df = pd.DataFrame(
        [
            [sapcode, interim_per_mat_dict[sapcode]["balance"]]
            for sapcode in interim_per_mat_dict.keys()
        ],
        columns=["Sapcode", "Balance"],
    )

    # Add in Mat Desc using lookup of material no(sapcode) from master_df
    balance_df["Material Description"] = balance_df.apply(
        lambda row: (
            (master_df.loc[master_df["Material Number"] == row["Sapcode"]])[
                "Material Description"
            ].to_list()
        )[0],
        axis=1,
    )
    og_balance_df_shape = balance_df.shape[0]

    if mode == "freestock":
        # Filter away balances < 1
        balance_df = balance_df.loc[balance_df["Balance"] >= 1]

        # Add in prod cat from material_to_cat mapping table
        amt1 = read_in_backend_excel(
            mw, "material_to_cat_mapping_table.xlsx", "Mapping Table", 7
        )
        balance_df = balance_df.merge(
            amt1[["Code", "Product Cat"]],
            how="left",
            left_on="Sapcode",
            right_on="Code",
        )
        balance_df = balance_df.drop("Code", axis=1)

        # Add in net weight from material_to_weight mapping table
        amt2 = read_in_backend_excel(
            mw, "material_to_weight_mapping_table.xlsx", "Mapping Table", 5
        )
        balance_df = balance_df.merge(
            amt2[["Material", "Net weight"]],
            how="left",
            left_on="Sapcode",
            right_on="Material",
        )
        balance_df = balance_df.drop("Material", axis=1)

        # Add in Total MT column
        balance_df["Total MT"] = (
            pd.to_numeric(balance_df["Balance"], errors="coerce")
            * pd.to_numeric(balance_df["Net weight"], errors="coerce")
            / 1000
        ).round(2)

        # Add comments of each sapcode from prev ecd check file
        # PENDING, COS COMMENTS HAS TO BE MERGED IN VIA PRI KEY, NOT SAPCODE

        # Pad rows with product cat NaN + create new grouped by balance_df aggregate sum of each product cat
        # Add in Est Pallets
        grouped_by_product_cat_balance_df = balance_df.copy()
        grouped_by_product_cat_balance_df["Product Cat"].fillna(
            "No Cat Found", inplace=True
        )
        grouped_by_product_cat_balance_df = (
            grouped_by_product_cat_balance_df.groupby("Product Cat")["Total MT"]
            .sum()
            .reset_index()
        )
        grouped_by_product_cat_balance_df["Est Pallets"] = (
            grouped_by_product_cat_balance_df["Total MT"] / 0.72
        ).round()

        new_row = {
            "Product Cat": "---Grand Total:---",
            "Total MT": round(grouped_by_product_cat_balance_df["Total MT"].sum(), 2),
            "Est Pallets": round(
                grouped_by_product_cat_balance_df["Est Pallets"].sum()
            ),
        }
        grouped_by_product_cat_balance_df = pd.concat(
            [grouped_by_product_cat_balance_df, pd.DataFrame([new_row])],
            ignore_index=True,
        )

    if mode == "main":
        for r in dataframe_to_rows(balance_df, index=False):
            myws.append(r)
    else:

        def add_dataframe_to_worksheet(df, start_row, start_col):
            for r_idx, row in enumerate(
                dataframe_to_rows(df, index=False, header=True), start_row
            ):
                for c_idx, value in enumerate(row, start_col):
                    myws.cell(row=r_idx, column=c_idx, value=value)
                if start_col > 1:
                    myws.cell(row=r_idx, column=start_col - 1, value="")

        add_dataframe_to_worksheet(balance_df, start_row=1, start_col=1)
        add_dataframe_to_worksheet(
            grouped_by_product_cat_balance_df, start_row=1, start_col=9
        )

    # Autospace all cells
    for column in myws.columns:
        max_length = max(len(str(cell.value)) for cell in column)
        column_letter = op.utils.get_column_letter(column[0].column)
        adjusted_width = max(max_length, 6) * 1.1
        myws.column_dimensions[column_letter].width = adjusted_width

    ## Shade column headers, shade all cells in first row if not empty, stop looking at col index 15
    for c_ind, column in enumerate(myws.iter_cols(min_row=1, max_row=1)):
        for cell in column:
            if cell.value is not None and cell.value != "":
                cell.fill = op.styles.PatternFill(
                    start_color="808080", end_color="808080", fill_type="solid"
                )
                cell.font = op.styles.Font(color="FFFFFF")
        if c_ind == 15:
            break
    return myws


def gr_generate_rupture(mw):
    """
    Actual generate ecd button perform operations
    Total of 8 Stages as of now, for progress tracking
    """
    mw.progress_window.displayPlease()
    try:
        # Verify all 4 files present
        all_generate_rupture_browse_fields_text = [
            txtfld.toPlainText() for txtfld in mw.gr_browse_text_fields
        ]
        if "" in all_generate_rupture_browse_fields_text:
            launch_message_box(
                mw,
                "Error: Missing Files",
                "All 4 files are required. You have not added one or more of the files, please check again",
            )
            return
        mw.progress_window.updateProgressBar(1 * 100 / 8)
        
        # Pandas read excel files, verify if each file correct col format, then store
        dict_of_pd_excel_dfs = {}
        for i, f in enumerate(single_source_of_truth.generate_rupture_files_required):
            file_path = mw.gr_browse_text_fields[i].toPlainText()
            df_from_excel = None
            if f == "Opening SOH":
                df_from_excel = pd.read_excel(
                    file_path,
                    sheet_name="Locally Produced",
                )
                # check until column 13 for opening SOH
                if not list(df_from_excel.columns)[:14] == single_source_of_truth.rupture_correct_columns[f]:
                    launch_message_box(
                        mw,
                        "Error: Invalid Excel Files Structure",
                        f"The excel file supplied for {f} has an invalid structure, please check",
                    )
                    return
            elif f == "ZC32":
                df_from_excel = pd.read_excel(
                    file_path
                )
                # check everything for zc32
                if not list(df_from_excel.columns) == single_source_of_truth.rupture_correct_columns[f]:
                    launch_message_box(
                        mw,
                        "Error: Invalid Excel Files Structure",
                        f"The excel file supplied for {f} has an invalid structure, please check",
                    )
                    return
            elif f == "Prod Vol":
                wb = op.load_workbook(file_path, read_only=True)
                sheet = wb["COOIS"]
                data = []
                for row in sheet.iter_rows(values_only=True):
                    data.append(row)
                df_from_excel = pd.DataFrame(data[1:], columns=data[0])
                
                # check first 22 cols only for prod vol
                if not list(df_from_excel.columns)[:23] == single_source_of_truth.rupture_correct_columns[f]:
                    launch_message_box(
                        mw,
                        "Error: Invalid Excel Files Structure",
                        f"The excel file supplied for {f} has an invalid structure, please check",
                    )
                    return
            elif f == "DN":
                df_from_excel = pd.read_excel(
                    file_path
                )
                # check everything for DN
                if not list(df_from_excel.columns) == single_source_of_truth.rupture_correct_columns[f]:
                    launch_message_box(
                        mw,
                        "Error: Invalid Excel Files Structure",
                        f"The excel file supplied for {f} has an invalid structure, please check",
                    )
                    return
            elif f == "Stock Transfer":
                df_from_excel = pd.read_excel(
                    file_path
                )
                # check everything for STO
                if not list(df_from_excel.columns) == single_source_of_truth.rupture_correct_columns[f]:
                    launch_message_box(
                        mw,
                        "Error: Invalid Excel Files Structure",
                        f"The excel file supplied for {f} has an invalid structure, please check",
                    )
                    return
            
            dict_of_pd_excel_dfs[f] = df_from_excel
        
        # If everything is ok, disable the app and start the process
        mw.central_widget.setDisabled(True)

        mw.progress_window.updateProgressBar(2 * 100 / 8)
        
        # Create one excel workbook obj
        wb = op.Workbook()
        # print(dict_of_pd_excel_dfs["DN"].columns)
        # print(dict_of_pd_excel_dfs["ZC32"]['Delivery Date'])
        # print(dict_of_pd_excel_dfs["ZC32"]['Reason For Rejection'])
        # return
        # ----------------- 1ST WORKSHEET: Rupture Consol ------------------------------
        newws = wb.active
        newws.title = "Rupture Consol"
        newws, interim_per_mat_dict, master_df = run_rupture_algorithm(
            mw, newws, dict_of_pd_excel_dfs
        )
        
        # Shade Nok status cells red
        red_fill = op.styles.PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")

        # Iterate through the 'Status' column and highlight 'Nok' cells
        for row in range(2, newws.max_row + 1):  # Start from row 2
            status_cell = newws.cell(row=row, column=9)  # 'Status' is in column 9
            if status_cell.value == 'Nok':
                status_cell.fill = red_fill  # Apply red fill to the cell
        
        # Autowidth spacing
        for column in newws.columns:
            max_length = max(len(str(cell.value)) for cell in column)
            column_letter = op.utils.get_column_letter(column[0].column)
            adjusted_width = max(max_length, 6) * 1.1
            newws.column_dimensions[column_letter].width = adjusted_width
        
        # Shade column headers, shade all cells in first row if not empty, stop looking at col index 17
        header_row_style = op.styles.NamedStyle(name="header_row_style")
        header_row_style.fill = op.styles.PatternFill(
            start_color="808080", end_color="808080", fill_type="solid"
        )
        header_row_style.font = op.styles.Font(color="FFFFFF")
        for c_ind, column in enumerate(newws.iter_cols(min_row=1, max_row=1)):
            for cell in column:
                if cell.value is not None and cell.value != "":
                    cell.style = header_row_style
            if c_ind == 15:
                break
        
        mw.progress_window.updateProgressBar(3 * 100 / 8)

        # # --------------- 2ND WORKSHEET: BALANCES FROM ECD -------------------------
        # # Create 2nd ws called 'Balances'
        # secws = wb.create_sheet("Balances")
        # secws = fill_sheet_with_balance_df(
        #     mw, secws, interim_per_mat_dict, master_df, mode="main"
        # )
        # mw.progress_window.updateProgressBar(4 * 100 / 8)

        # # ----------------- 3RD WORKSHEET: DN > SOH ------------------------------
        # # Calculate overshot between them
        # thirdws = wb.create_sheet("DN > SOH")
        # for r in range(2, master_df.shape[0] + 2):
        #     for sapcode in interim_per_mat_dict.keys():
        #         if sapcode == newws.cell(row=r, column=1).value:
        #             if (
        #                 newws.cell(row=r, column=6).value == "SOH"
        #                 or newws.cell(row=r, column=6).value == "DN"
        #             ):
        #                 interim_per_mat_dict[sapcode]["overshot"] += newws.cell(
        #                     row=r, column=9
        #                 ).value
        #             break
        # ## Create overshot df, add mat desc, get rid of positive counters
        # overshotdf = pd.DataFrame(
        #     [
        #         [sapcode, interim_per_mat_dict[sapcode]["overshot"]]
        #         for sapcode in interim_per_mat_dict.keys()
        #     ],
        #     columns=["Sapcode", "Counter"],
        # )
        # overshotdf["Material Description"] = overshotdf.apply(
        #     lambda row: (
        #         (master_df.loc[master_df["Material Number"] == row["Sapcode"]])[
        #             "Material Description"
        #         ].to_list()
        #     )[0],
        #     axis=1,
        # )
        # overshotdf = overshotdf.loc[overshotdf["Counter"] < 0]
        # for r in dataframe_to_rows(overshotdf, index=False):
        #     thirdws.append(r)
        # mw.progress_window.updateProgressBar(5 * 100 / 8)
        # # ----------------- 4TH WORKSHEET: SO DN ---------------------------------
        # fifthws = wb.create_sheet("SO DN")
        # ecddata = newws.values
        # ecdcolumns = next(ecddata)
        # ecddf = pd.DataFrame(ecddata, columns=ecdcolumns)

        # ## Function to extract only numeric values from 'SOH: ' inside ECD string
        # def extract_numeric_values(text):
        #     pattern = r"SOH: (\d+)"
        #     matches = re.findall(pattern, text)
        #     numeric_sum = sum(int(value) for value in matches)
        #     return numeric_sum

        # ## Iterate thru each type and aggregate them accordingly
        # agg_tables = []
        # amt2 = read_in_backend_excel(
        #     mw, "material_to_weight_mapping_table.xlsx", "Mapping Table", 5
        # )
        # for agg_name in single_source_of_truth.aggregate_table_names:
        #     type_mask = ecddf["Type"].isin(agg_name.split("/"))
        #     comments_mask = ecddf["ECD"].str.contains("SOH")
        #     ecddf_typed = ecddf[type_mask & comments_mask]
        #     ecddf_typed["Orders"] = ecddf_typed["ECD"].apply(extract_numeric_values)

        #     ### Read in amt, left join on mat no
        #     carmen = ecddf_typed.merge(
        #         amt2, how="left", left_on="Material Number", right_on="Material"
        #     )

        #     #### Retrieve desired columns only, calc total weight per row and convert to MT,pallets
        #     carmen = carmen[["Material Number", "Name", "Orders", "Net weight"]]
        #     carmen["Net weight"] = pd.to_numeric(carmen["Net weight"], errors="coerce")
        #     carmen["Net weight"].fillna(0, inplace=True)
        #     carmen["total_weight"] = carmen["Orders"] * carmen["Net weight"]
        #     carmen["Total MT"] = carmen["total_weight"] / 1000
        #     carmen["Estimated Pallets"] = carmen["Total MT"] / 0.72

        #     #### Rename Name to agg_name
        #     carmen = carmen.rename(columns={"Name": agg_name})

        #     #### Now group by Name, sum up Total MT and Estimated Pallets
        #     mt_sum = (
        #         carmen.groupby(agg_name)["Total MT"]
        #         .agg("sum")
        #         .reset_index()
        #         .sort_values(by="Total MT", ascending=False)
        #     )
        #     pallet_sum = (
        #         carmen.groupby(agg_name)["Estimated Pallets"]
        #         .agg("sum")
        #         .reset_index()
        #         .sort_values(by="Estimated Pallets", ascending=False)
        #     )

        #     #### join mt_sum and pallet_sum into one table + convert to 2dp & whole no
        #     joined_mt_pallet_sum = pd.merge(mt_sum, pallet_sum, on=agg_name)
        #     joined_mt_pallet_sum["Total MT"] = joined_mt_pallet_sum["Total MT"].round(2)
        #     joined_mt_pallet_sum["Estimated Pallets"] = joined_mt_pallet_sum[
        #         "Estimated Pallets"
        #     ].round()

        #     #### Add tabulation last row
        #     new_row = {
        #         agg_name: "-----TOTAL SUMS:-----",
        #         "Total MT": joined_mt_pallet_sum["Total MT"].sum(),
        #         "Estimated Pallets": joined_mt_pallet_sum["Estimated Pallets"].sum(),
        #     }
        #     joined_mt_pallet_sum = pd.concat(
        #         [joined_mt_pallet_sum, pd.DataFrame([new_row])], ignore_index=True
        #     )

        #     #### Store ready aggregated dfs into agg_tables list
        #     agg_tables.append(joined_mt_pallet_sum)

        # ## Write to fifthws
        # max_rows = max(
        #     len(list(dataframe_to_rows(df, index=False, header=True)))
        #     for df in agg_tables
        # )
        # for i in range(max_rows):
        #     combined_row = []
        #     for df in agg_tables:
        #         rows = list(dataframe_to_rows(df, index=False, header=True))
        #         if i < len(rows):
        #             combined_row.extend(rows[i])
        #         else:
        #             combined_row.extend([None] * len(rows[0]))
        #         combined_row.append(None)
        #     fifthws.append(combined_row)

        # ## Shade column headers, shade all cells in first row if not empty, stop looking at col index 15
        # header_row_style = op.styles.NamedStyle(name="header_row_style")
        # header_row_style.fill = op.styles.PatternFill(
        #     start_color="808080", end_color="808080", fill_type="solid"
        # )
        # header_row_style.font = op.styles.Font(color="FFFFFF")
        # for c_ind, column in enumerate(fifthws.iter_cols(min_row=1, max_row=1)):
        #     for cell in column:
        #         if cell.value is not None and cell.value != "":
        #             cell.style = header_row_style
        #     if c_ind == 15:
        #         break

        # ## Autowidth spacing and left justify for fifthws
        # for column in fifthws.columns:
        #     max_length = max(len(str(cell.value)) for cell in column)
        #     column_letter = op.utils.get_column_letter(column[0].column)
        #     adjusted_width = max(max_length, 6) * 1.1
        #     fifthws.column_dimensions[column_letter].width = adjusted_width
        # mw.progress_window.updateProgressBar(6 * 100 / 8)

        # # ---------------- 5TH WORKSHEET: FREE STOCK BALANCES FROM ECD W/O OSIL -----------------------------
        # # Basically, rerun the main ecd algo w/o OISL mode freestock and get balances agn like 2nd ws
        # interimws = wb.create_sheet("interim")
        # seventhws = wb.create_sheet("free stock")
        # (
        #     interimws,
        #     interim_per_mat_dict_free_stock,
        #     master_df_free_stock,
        # ) = run_ecd_algorithm(mw, interimws, dict_of_pd_excel_dfs, mode="freestock")
        # seventhws = fill_sheet_with_balance_df(
        #     mw,
        #     seventhws,
        #     interim_per_mat_dict_free_stock,
        #     master_df_free_stock,
        #     mode="freestock",
        # )
        # wb.remove(wb["interim"])
        # mw.progress_window.updateProgressBar(7 * 100 / 8)

        # # ---------------- 6TH WORKSHEET: ZC32 CS LAYOUT -------------------------
        # forthws = wb.create_sheet("zc32 cs layout")

        # ## Drop last tabulation row
        # dict_of_pd_excel_dfs["ZC32CS"].drop(
        #     dict_of_pd_excel_dfs["ZC32CS"].index[-1], inplace=True
        # )

        # ## Create Primary Key & P doc column and put to front
        # dict_of_pd_excel_dfs["ZC32CS"]["Sales Document"] = pd.to_numeric(
        #     dict_of_pd_excel_dfs["ZC32CS"]["Sales Document"],
        #     errors="coerce",
        #     downcast="integer",
        # )
        # dict_of_pd_excel_dfs["ZC32CS"]["Item No"] = pd.to_numeric(
        #     dict_of_pd_excel_dfs["ZC32CS"]["Item No"],
        #     errors="coerce",
        #     downcast="integer",
        # )
        # dict_of_pd_excel_dfs["ZC32CS"]["Primary Key"] = dict_of_pd_excel_dfs["ZC32CS"][
        #     "Sales Document"
        # ].astype(str) + dict_of_pd_excel_dfs["ZC32CS"]["Item No"].astype(str)
        # dict_of_pd_excel_dfs["ZC32CS"]["Purchasing Document"] = ""
        # zcolumns = list(dict_of_pd_excel_dfs["ZC32CS"].columns)
        # zcolumns.insert(0, zcolumns.pop(zcolumns.index("Primary Key")))
        # zcolumns.insert(1, zcolumns.pop(zcolumns.index("Purchasing Document")))
        # dict_of_pd_excel_dfs["ZC32CS"] = dict_of_pd_excel_dfs["ZC32CS"][zcolumns]

        # ## Cast open qty to int, truncate decimal places
        # dict_of_pd_excel_dfs["ZC32CS"]["Open Qty."] = dict_of_pd_excel_dfs["ZC32CS"][
        #     "Open Qty."
        # ].astype(int)

        # ## Remove non-blank Reason For Rejection Rows
        # dict_of_pd_excel_dfs["ZC32CS"] = dict_of_pd_excel_dfs["ZC32CS"][
        #     pd.isna(dict_of_pd_excel_dfs["ZC32CS"]["Reason For Rejection"])
        # ]

        # ## Drop Reason For Rejection column
        # dict_of_pd_excel_dfs["ZC32CS"] = dict_of_pd_excel_dfs["ZC32CS"].drop(
        #     "Reason For Rejection", axis=1
        # )

        # ## left join zc32 cs layout twice with both ecd df and prev ecd df to get ecd status and comments
        # ecddata = newws.values
        # ecdcolumns = next(ecddata)
        # ecddf = pd.DataFrame(ecddata, columns=ecdcolumns)
        # dict_of_pd_excel_dfs["ZC32CS"] = dict_of_pd_excel_dfs["ZC32CS"].merge(
        #     ecddf[["Primary Key", "ECD"]], on="Primary Key", how="left"
        # )
        # dict_of_pd_excel_dfs["Prev ECD"]["Primary Key"] = dict_of_pd_excel_dfs[
        #     "Prev ECD"
        # ]["Primary Key"].astype(str)
        # dict_of_pd_excel_dfs["ZC32CS"] = dict_of_pd_excel_dfs["ZC32CS"].merge(
        #     dict_of_pd_excel_dfs["Prev ECD"][["Primary Key", "Comments"]],
        #     on="Primary Key",
        #     how="left",
        # )
        # dict_of_pd_excel_dfs["ZC32CS"].rename(
        #     columns={"ECD": "ECD Status"}, inplace=True
        # )

        # ## Sort by ascending tuple (mat no, first delivery date)
        # dict_of_pd_excel_dfs["ZC32CS"] = dict_of_pd_excel_dfs["ZC32CS"].sort_values(
        #     by=["Material Number", "First Delivery date"], ascending=True
        # )

        # ## Cast all date columns to raw date formats for parsing back into excel
        # dict_of_pd_excel_dfs["ZC32CS"]["Doc. Date"] = pd.to_datetime(
        #     dict_of_pd_excel_dfs["ZC32CS"]["Doc. Date"]
        # )
        # dict_of_pd_excel_dfs["ZC32CS"]["First Delivery date"] = pd.to_datetime(
        #     dict_of_pd_excel_dfs["ZC32CS"]["First Delivery date"]
        # )

        # ## Write dataframe to forth excel sheet
        # for r in dataframe_to_rows(dict_of_pd_excel_dfs["ZC32CS"], index=False):
        #     forthws.append(r)

        # ## Apply date formatting style in forthws
        # date_style = op.styles.NamedStyle(name="date_style", number_format="DD/MM/YYYY")
        # for cell in forthws["E"][1:]:
        #     cell.style = date_style
        # for cell in forthws["R"][1:]:
        #     cell.style = date_style

        # ## Apply numeric formatting style in forthws
        # numeric_format_0dp = op.styles.NamedStyle(
        #     name="numeric_format", number_format="0"
        # )
        # numeric_format_2dp = op.styles.NamedStyle(
        #     name="numeric_format1", number_format="0.00"
        # )
        # numeric_format_3dp = op.styles.NamedStyle(
        #     name="numeric_format2", number_format="0.000"
        # )
        # for cell in forthws["N"][1:]:
        #     cell.style = numeric_format_0dp
        # for cell in forthws["O"][1:]:
        #     cell.style = numeric_format_3dp
        # for cell in forthws["P"][1:]:
        #     cell.style = numeric_format_2dp
        # for cell in forthws["Q"][1:]:
        #     cell.style = numeric_format_2dp

        # ## Color this tab yellow
        # forthws.sheet_properties.tabColor = op.styles.Color(rgb="FFFF00")

        # ## Color header row to grey, orange for certain columns
        # grey_fill = op.styles.PatternFill(
        #     start_color="C0C0C0", end_color="C0C0C0", fill_type="solid"
        # )
        # orange_fill = op.styles.PatternFill(
        #     start_color="FFA500", end_color="FFA500", fill_type="solid"
        # )
        # for col_idx, col in enumerate(dict_of_pd_excel_dfs["ZC32CS"].columns, start=1):
        #     cell = forthws.cell(row=1, column=col_idx)
        #     if col in single_source_of_truth.zc32_sheet_orange_columns:
        #         cell.fill = orange_fill
        #     else:
        #         cell.fill = grey_fill

        # ## Set left justify & autowidth spacing, except Comments column index 22
        # col13width = -99
        # for column in forthws.columns:
        #     # Column 22 (comments) do not autowidth
        #     if column[0].column == 22:
        #         pass
        #     elif column[0].column != 21:
        #         # The rest of the cols except column 21, autowidth
        #         max_length = max(len(str(cell.value)) for cell in column)
        #         column_letter = op.utils.get_column_letter(column[0].column)
        #         adjusted_width = max(max_length, 6) * 1.1
        #         forthws.column_dimensions[column_letter].width = adjusted_width

        #         # Column 13 (mat desc) record down its width
        #         if column[0].column == 13:
        #             col13width = adjusted_width

        # # Column 21 (ecd status) use col13 width
        # forthws.column_dimensions["U"].width = col13width

        # Save the excel workbook into dir and close it, if not autofind then prompt savewindow else save to perma working dir using ecdgen name
        if mw.ss_perma_working_dir_text_field.toPlainText() == "":
            savepath, _ = QFileDialog.getSaveFileName(
                QWidget(), "Choose a place to generate to", "", "Excel files (*.xlsx)"
            )
            if savepath == "":
                launch_message_box(
                    mw,
                    "Error: Unable to save generated file",
                    "Directory to save file in cannot be blank, please try generating again",
                )
                return
            else:
                wb.save(savepath)
        else:
            default_rupture_gen_name = mw.ss_ecdgen_filename_text_field.text()
            if default_rupture_gen_name == "":
                launch_message_box(
                    mw,
                    "Error: Unable to save generated file",
                    "Default rupture consol report filename not set. Please head to settings to specify default generated filename eg: somegeneratedfile",
                )
                return
            wb.save(
                mw.ss_perma_working_dir_text_field.toPlainText()
                + "/"
                + default_rupture_gen_name
                + ".xlsx"
            )

        wb.close()

        # Renable app of course and prompt complete using progress box widget
        mw.central_widget.setDisabled(False)
        mw.progress_window.updateInternalText("Rupture Consol Report file has been generated.")
        mw.progress_window.updateProgressBar(8 * 100 / 8)
        mw.progress_window.unhideOkButton()

    except Exception as ee:
        exception_name = type(ee).__name__
        exception_description = str(ee)
        launch_message_box(
            mw,
            "An Error has occurred, please report incident to Marcus thanks",
            f"{exception_name},{exception_description}",
        )
        traceback.print_exc()
        mw.central_widget.setDisabled(False)


def init_gui_behaviour(mw):
    """Sets basic behaviours for the mainwindow"""
    # Set starting page
    mw.central_stack_of_pages.setCurrentIndex(0)

    # On sidebar button press, go to page
    mw.generate_rupture_button.clicked.connect(
        lambda: mw.central_stack_of_pages.setCurrentIndex(0)
    )
    mw.vbsloader_button.clicked.connect(
        lambda: mw.central_stack_of_pages.setCurrentIndex(1)
    )
    mw.settings_button.clicked.connect(
        lambda: mw.central_stack_of_pages.setCurrentIndex(2)
    )

    # Generate rupture page buttons behaviours
    for ind, mb in enumerate(mw.list_of_menus):
        if mb == "Auto Find Files":
            mw.gr_menu_buttons[ind].clicked.connect(lambda: gr_auto_find_all_files(mw))
        elif mb == "Clear All":
            mw.gr_menu_buttons[ind].clicked.connect(
                lambda: gr_clear_all_added_files(mw)
            )

    for ind, ft in enumerate(mw.list_of_files_to_browse):
        if ft == "Opening SOH":
            mw.gr_browse_buttons[ind].clicked.connect(
                lambda: gr_add_browsed_file(mw, "Opening SOH")
            )
        elif ft == "ZC32":
            mw.gr_browse_buttons[ind].clicked.connect(
                lambda: gr_add_browsed_file(mw, "ZC32")
            )
        elif ft == "Prod Vol":
            mw.gr_browse_buttons[ind].clicked.connect(
                lambda: gr_add_browsed_file(mw, "Prod Vol")
            )
        elif ft == "DN":
            mw.gr_browse_buttons[ind].clicked.connect(
                lambda: gr_add_browsed_file(mw, "DN")
            )

    mw.gr_generate_rupture_button.clicked.connect(lambda: gr_generate_rupture(mw))

    # Settings page buttons
    for inf, acb in enumerate(mw.list_of_accent_colors):
        mw.ss_accent_radio_buttons[inf].toggled.connect(
            lambda: ss_change_ascent_color(mw)
        )
    mw.ss_browse_permadir_button.clicked.connect(
        lambda: ss_change_perma_working_dir(mw)
    )
    mw.ss_clear_permadir_button.clicked.connect(lambda: ss_clear_perma_working_dir(mw))
    mw.ss_ecdgen_filename_text_field.editingFinished.connect(
        lambda: (
            mw.ss_ecdgen_filename_text_field.setText(
                mw.ss_ecdgen_filename_text_field.text()
            ),
            mw.ss_ecdgen_filename_text_field.clearFocus(),
            ss_change_default_ecdgen_name(mw),
        )
    )

    # vbs loader page buttons behaviours

    ### vb_menu_buttons[0] --> browse vbs buton
    ### vb_menu_buttons[1] --> clear all vbs buton
    mw.vb_menu_buttons[0].clicked.connect(
        lambda: vb_add_browsed_vbs_file(mw, "VBscript")
    )
    mw.vb_menu_buttons[1].clicked.connect(lambda: vb_clear_all_vbs_history_list(mw))

    mw.vb_runvbs_button.clicked.connect(lambda: vb_execute_vbs_script(mw))
